import logging
from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QGroupBox, QScrollArea, 
    QTableWidget, QTableWidgetItem, QGridLayout, QHeaderView, QPushButton
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
import openpyxl
import os

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

class CurvedPerformanceView(QWidget):
    def __init__(self, parent_app):
        super().__init__()
        logging.debug("Initializing CurvedPerformanceView")
        self.parent_app = parent_app
        self.setMinimumSize(1600, 1200)
        self.setStyleSheet("""
            QWidget { 
                background-color: #f5f5f5; 
                font-family: 'Segoe UI', sans-serif; 
            }
            QGroupBox { 
                font-weight: bold; 
                font-size: 12px; 
                border: 2px solid #cccccc; 
                border-radius: 5px; 
                margin-top: 10px; 
                padding-top: 10px; 
                background-color: white; 
            }
            QGroupBox::title { 
                subcontrol-origin: margin; 
                left: 10px; 
                padding: 0 5px; 
            }
            QTableWidget { 
                border: 1px solid #cccccc; 
                background-color: white;
                gridline-color: #cccccc;
                font-size: 10px;
            }
            QTableWidget::item {
                padding: 3px;
                border: 1px solid #cccccc;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #e8e8e8;
                padding: 5px;
                border: 1px solid #cccccc;
                font-weight: bold;
            }
        """)
        self.create_ui()
        # logging.debug("UI created")
        if not self.load_data():
            # logging.error("Failed to load performance data")
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Error", "Failed to load performance data. Please check the Excel file.")
            if self.parent_app:
                self.parent_app.go_back()
        else:
            # logging.debug("Data loaded, scheduling curves update")
            QTimer.singleShot(0, self.update_curves)

    def create_ui(self):
        # logging.debug("Creating UI")
        main_layout = QVBoxLayout(self)
        
        # Add back button
        back_button = QPushButton("Back to Form")
        back_button.setStyleSheet("""
            QPushButton {
                background-color: #212580;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #3439C7;
            }
        """)
        if self.parent_app:
            back_button.clicked.connect(self.parent_app.go_back)
        main_layout.addWidget(back_button)
        
        header_label = QLabel("Annual Performance Review 2024/2025\n(Performance Curve)")
        header_label.setAlignment(Qt.AlignCenter)
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("background-color: #2c3e50; color: white; padding: 15px; border-radius: 5px;")
        main_layout.addWidget(header_label)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        company_group = QGroupBox("Company Information")
        company_layout = QGridLayout(company_group)
        dept_label = QLabel("Department/Company Name:")
        dept_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.dept_value = QLabel("Pakistan Machine Tool Factory")
        self.dept_value.setStyleSheet("background-color: #e8f4f8; padding: 5px; border: 1px solid #ccc;")
        company_layout.addWidget(dept_label, 0, 0)
        company_layout.addWidget(self.dept_value, 0, 1)
        emp_count_label = QLabel("Total Employee Count:")
        emp_count_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.emp_count_value = QLabel("0")
        self.emp_count_value.setStyleSheet("background-color: #ffff99; padding: 5px; border: 1px solid #ccc; font-weight: bold;")
        company_layout.addWidget(emp_count_label, 1, 0)
        company_layout.addWidget(self.emp_count_value, 1, 1)
        scroll_layout.addWidget(company_group)
        summary_group = QGroupBox("Performance Distribution Summary")
        summary_layout = QVBoxLayout(summary_group)
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(6)
        self.summary_table.setHorizontalHeaderLabels([
            "Performance Rating", "Performance Curve", "Required Distribution of Employees", 
            "Actual Distribution after Rating", "Actual Performance Curve", "Percentage"
        ])
        self.summary_table.setRowCount(6)  
        self.summary_table.horizontalHeader().setStretchLastSection(True)
        summary_layout.addWidget(self.summary_table)
        scroll_layout.addWidget(summary_group)
        charts_group = QGroupBox("Performance Curve Visualization")
        charts_layout = QHBoxLayout(charts_group)
        actual_chart_layout = QVBoxLayout()
        actual_title = QLabel("Actual Performance Curve")
        actual_title.setAlignment(Qt.AlignCenter)
        actual_title.setFont(QFont("Arial", 12, QFont.Bold))
        actual_chart_layout.addWidget(actual_title)
        self.actual_chart = QChart()
        self.actual_chart_view = QChartView(self.actual_chart)
        self.actual_chart_view.setMinimumSize(400, 300)
        actual_chart_layout.addWidget(self.actual_chart_view)
        charts_layout.addLayout(actual_chart_layout)
        required_chart_layout = QVBoxLayout()
        required_title = QLabel("Required Performance Curve")
        required_title.setAlignment(Qt.AlignCenter)
        required_title.setFont(QFont("Arial", 12, QFont.Bold))
        required_chart_layout.addWidget(required_title)
        self.required_chart = QChart()
        self.required_chart_view = QChartView(self.required_chart)
        self.required_chart_view.setMinimumSize(400, 300)
        required_chart_layout.addWidget(self.required_chart_view)
        charts_layout.addLayout(required_chart_layout)
        scroll_layout.addWidget(charts_group)
        employee_group = QGroupBox("Employee Performance Data")
        employee_layout = QVBoxLayout(employee_group)
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(11)
        self.employee_table.setHorizontalHeaderLabels([
            "Emp No", "Employee Name", "Designation", "Tier", "Company", 
            "Division / Department", "Employment Category", "Employment Type", "DOJ", "Rating (Sample)", "Scores"
        ])
        header = self.employee_table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.Stretch) 
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)  
        self.employee_table.setMinimumHeight(300)
        employee_layout.addWidget(self.employee_table)
        scroll_layout.addWidget(employee_group)
        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll)
        # logging.debug("UI creation complete")
        

    def load_data(self):
        try:
            # logging.debug("Loading data from employee_performance_data.xlsx")
            emp_file = os.path.join("assets", "employee_performance_data.xlsx")
            if not os.path.exists(emp_file):
                logging.error(f"Employee file not found: {emp_file}")
                return False
            wb_emp = openpyxl.load_workbook(emp_file)
            sheet_names = wb_emp.sheetnames
            ws_emp = None
            for sheet_name in ["Sheet1", "Performance_Data", "Data", "Employee_Data"]:
                if sheet_name in sheet_names:
                    ws_emp = wb_emp[sheet_name]
                    break
            if not ws_emp:
                logging.error(f"No valid sheet found. Available sheets: {sheet_names}")
                return False
            self.employee_data = []
            headers = []
            for cell in ws_emp[1]:
                headers.append(cell.value)
            # logging.debug(f"Headers found: {headers[:15]}...")  
            employee_id_col = self.find_column_index(headers, ["Employee ID", "Emp ID", "ID"])
            employee_name_col = self.find_column_index(headers, ["Employee Name", "Name"])
            department_col = self.find_column_index(headers, ["Department", "Dept"])
            designation_col = self.find_column_index(headers, ["Designation", "Position"])
            doj_col = self.find_column_index(headers, ["Date of Joining", "DOJ", "Joining Date"])
            division_col = self.find_column_index(headers, ["Division", "Div"])
            overall_rating_col = self.find_column_index(headers, ["Overall_Rating", "Overall Rating", "Rating"])
            overall_percentage_col = self.find_column_index(headers, ["Overall_Percentage", "Overall Percentage", "Percentage"])
            
            for row in ws_emp.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 0 and row[0]:  # Employee ID exists
                    employee_id = row[employee_id_col] if employee_id_col is not None else row[0]
                    employee_name = row[employee_name_col] if employee_name_col is not None else (row[1] if len(row) > 1 else "")
                    department = row[department_col] if department_col is not None else (row[2] if len(row) > 2 else "")
                    designation = row[designation_col] if designation_col is not None else (row[3] if len(row) > 3 else "")
                    doj = row[doj_col] if doj_col is not None else (row[4] if len(row) > 4 else "")
                    division = row[division_col] if division_col is not None else (row[6] if len(row) > 6 else "")
                    overall_rating = row[overall_rating_col] if overall_rating_col is not None else (
                        row[43] if len(row) > 43 else "Meets Expectations (3)")
                    overall_percentage = row[overall_percentage_col] if overall_percentage_col is not None else (
                        row[44] if len(row) > 44 else 0)
                    
                    if isinstance(overall_percentage, str) and '%' in str(overall_percentage):
                        overall_percentage = float(str(overall_percentage).replace('%', ''))
                    elif overall_percentage is None:
                        overall_percentage = 0
                    
                    self.employee_data.append({
                        "id": str(employee_id) if employee_id else "",
                        "name": str(employee_name) if employee_name else "",
                        "department": str(department) if department else "",
                        "designation": str(designation) if designation else "",
                        "doj": str(doj) if doj else "",
                        "division": str(division) if division else "",
                        "overall_rating": str(overall_rating) if overall_rating else "Meets Expectations (3)",
                        "overall_percentage": float(overall_percentage) if overall_percentage else 0
                    })
            
            # logging.debug(f"Loaded {len(self.employee_data)} employee records")
            if not self.employee_data:
                logging.error("No employee data loaded")
                return False
            return True
        except Exception as e:
            logging.error(f"Error loading data: {str(e)}")
            return False

    def find_column_index(self, headers, possible_names):
        for i, header in enumerate(headers):
            if header:
                header_str = str(header).strip()
                for name in possible_names:
                    if name.lower() in header_str.lower():
                        return i
        return None

    def update_curves(self):
        # logging.debug("Starting update_curves")
        try:
            ratings = [
                "Outstanding (5)", 
                "Exceeds Expectations (4)",
                "Meets Expectations (3)", 
                "Below Expectations (2)", 
                "Serious Performance Concerns (1)"
            ]
            
            required_percentages = {
                "Outstanding (5)": 5,
                "Exceeds Expectations (4)": 15,
                "Meets Expectations (3)": 65,
                "Below Expectations (2)": 10,
                "Serious Performance Concerns (1)": 5
            }
            actual_counts = {rating: 0 for rating in ratings}
            # logging.debug("Sample employee ratings:")
            for i, emp in enumerate(self.employee_data[:10]):  
                logging.debug(f"Employee {i+1}: Name='{emp.get('name', 'No name')}', Rating='{emp.get('overall_rating', 'No rating')}'")
            for emp in self.employee_data:
                rating = emp.get("overall_rating", "")
                if rating: 
                    rating_str = str(rating).strip().lower()
                    if "5" in rating_str or "outstanding" in rating_str:
                        actual_counts["Outstanding (5)"] += 1
                    elif "4" in rating_str or "exceed" in rating_str:
                        actual_counts["Exceeds Expectations (4)"] += 1
                    elif "2" in rating_str or "below" in rating_str:
                        actual_counts["Below Expectations (2)"] += 1
                    elif "1" in rating_str or "serious" in rating_str or "poor" in rating_str:
                        actual_counts["Serious Performance Concerns (1)"] += 1
                    else:
                        actual_counts["Meets Expectations (3)"] += 1
                else:
                    actual_counts["Meets Expectations (3)"] += 1
            
            total_employees = len(self.employee_data)
            self.emp_count_value.setText(str(total_employees))
            # logging.debug(f"REAL-TIME Actual counts: {actual_counts}")
            # logging.debug(f"Total employees from data: {total_employees}")
            total_counted = sum(actual_counts.values())
            if total_counted != total_employees:
                logging.warning(f"Count mismatch! Total employees: {total_employees}, Total counted: {total_counted}")
            self.update_summary_table(ratings, required_percentages, actual_counts, total_employees)
            self.create_actual_pie_chart(actual_counts, total_employees)
            self.create_required_pie_chart(required_percentages)
            self.update_employee_table()
        except Exception as e:
            logging.error(f"Error in update_curves: {str(e)}")
            import traceback
            traceback.print_exc()

    def update_summary_table(self, ratings, required_percentages, actual_counts, total_employees):
        rating_data = [
            ("Outstanding (5)", 5),
            ("Exceeds Expectations(4)", 15), 
            ("Meets Expectations (3)", 65),
            ("Below Expectations (2)", 10),
            ("Serious Performance Concerns (1)", 5)
        ]
        
        total_required = 0
        total_actual = 0
        for i, (rating, req_percent) in enumerate(rating_data):
            item = QTableWidgetItem(rating)
            self.summary_table.setItem(i, 0, item)
            item = QTableWidgetItem(f"{req_percent}%")
            item.setTextAlignment(Qt.AlignCenter)
            self.summary_table.setItem(i, 1, item)
            required_counts = [3, 9, 38, 6, 3]
            req_count = required_counts[i]
            item = QTableWidgetItem(str(req_count))
            item.setTextAlignment(Qt.AlignCenter)
            self.summary_table.setItem(i, 2, item)
            total_required += req_count
            if i == 0:  
                actual_count = actual_counts.get("Outstanding (5)", 0)
            elif i == 1:  
                actual_count = actual_counts.get("Exceeds Expectations (4)", 0)
            elif i == 2:  
                actual_count = actual_counts.get("Meets Expectations (3)", 0)
            elif i == 3:  
                actual_count = actual_counts.get("Below Expectations (2)", 0)
            elif i == 4:  
                actual_count = actual_counts.get("Serious Performance Concerns (1)", 0)
            else:
                actual_count = 0
            item = QTableWidgetItem(str(actual_count))
            item.setTextAlignment(Qt.AlignCenter)
            item.setBackground(QColor("#e8f5e8"))  
            self.summary_table.setItem(i, 3, item)
            total_actual += actual_count
            actual_percent = (actual_count / total_employees * 100) if total_employees > 0 else 0
            item = QTableWidgetItem(f"{actual_percent:.1f}%") 
            item.setTextAlignment(Qt.AlignCenter)
            item.setBackground(QColor("#e8f5e8"))  
            self.summary_table.setItem(i, 4, item)
            item = QTableWidgetItem(f"{actual_percent:.1f}%")
            item.setTextAlignment(Qt.AlignCenter)
            item.setBackground(QColor("#e8f5e8")) 
            self.summary_table.setItem(i, 5, item)
        
        total_row = 5
        total_required_calc = sum([3, 9, 38, 6, 3])   
        total_actual_calc = total_actual
        
        total_items = [
            ("Total", 0), 
            ("100%", 1), 
            (str(total_required_calc), 2), 
            (str(total_actual_calc), 3), 
            (f"{(total_actual_calc/total_employees*100):.1f}%" if total_employees > 0 else "0%", 4), 
            (f"{(total_actual_calc/total_employees*100):.1f}%" if total_employees > 0 else "0%", 5)
        ]
        
        for text, col in total_items:
            item = QTableWidgetItem(text)
            item.setTextAlignment(Qt.AlignCenter)
            font = item.font()
            font.setBold(True)
            item.setFont(font)
            if col >= 3:  
                item.setBackground(QColor("#d4e6d4"))  
            else:
                item.setBackground(QColor("#e8e8e8"))
            self.summary_table.setItem(total_row, col, item)
        
        self.summary_table.resizeColumnsToContents()
        
        # logging.debug(f"REAL-TIME Summary table updated:")
        # logging.debug(f"  Total Required: {total_required_calc}")
        # logging.debug(f"  Total Actual: {total_actual_calc}")
        for rating, count in actual_counts.items():
            logging.debug(f"  {rating}: {count} employees")

    def create_actual_pie_chart(self, actual_counts, total_employees):
        self.actual_chart.removeAllSeries()
        series = QPieSeries()
        
        colors = ["#70ad47", "#ffc000", "#5b9bd5", "#ff9933", "#c55a5a"]
        
        chart_data = [
            ("Outstanding (5)", actual_counts.get("Outstanding (5)", 0)),
            ("Exceeds Expectations (4)", actual_counts.get("Exceeds Expectations (4)", 0)),
            ("Meets Expectations (3)", actual_counts.get("Meets Expectations (3)", 0)),
            ("Below Expectations (2)", actual_counts.get("Below Expectations (2)", 0)),
            ("Serious Performance Concerns (1)", actual_counts.get("Serious Performance Concerns (1)", 0))
        ]
        for i, (rating, count) in enumerate(chart_data):
            percentage = (count / total_employees * 100) if total_employees > 0 else 0
            if count > 0:
                label = rating.split('(')[0].strip()
                slice_obj = QPieSlice(f"{label}\n{percentage:.0f}%", percentage)
                slice_obj.setBrush(QColor(colors[i % len(colors)]))
                slice_obj.setLabelVisible(True)
                slice_obj.setExploded(False)
                series.append(slice_obj)
        if total_employees == 0 or sum(actual_counts.values()) == 0:
            slice_obj = QPieSlice("No Data\n0%", 100)
            slice_obj.setBrush(QColor("#cccccc"))
            slice_obj.setLabelVisible(True)
            series.append(slice_obj)
        self.actual_chart.addSeries(series)
        self.actual_chart.setTitle("")
        self.actual_chart.legend().setVisible(False)
        self.actual_chart_view.update()

    def create_required_pie_chart(self, required_percentages):
        """Create pie chart for required distribution matching Excel"""
        self.required_chart.removeAllSeries()
        series = QPieSeries()
        colors = ["#70ad47", "#ffc000", "#5b9bd5", "#ff9933", "#c55a5a"]
        chart_data = [
            ("Outstanding (5)", 5),
            ("Exceeds Expectations (4)", 15),
            ("Meets Expectations (3)", 65),
            ("Below Expectations (2)", 10),
            ("Serious Performance Concerns (1)", 5)
        ]
        for i, (rating, percentage) in enumerate(chart_data):
            label = rating.split('(')[0].strip()
            slice_obj = QPieSlice(f"{label}\n{percentage}%", percentage)
            slice_obj.setBrush(QColor(colors[i % len(colors)]))
            slice_obj.setLabelVisible(True)
            slice_obj.setExploded(False)
            series.append(slice_obj)
        self.required_chart.addSeries(series)
        self.required_chart.setTitle("")
        self.required_chart.legend().setVisible(False)
        self.required_chart_view.update()

    def update_employee_table(self):
        self.employee_table.setRowCount(len(self.employee_data))
        for row, emp in enumerate(self.employee_data):
            self.employee_table.setItem(row, 0, QTableWidgetItem(str(emp["id"])))
            self.employee_table.setItem(row, 1, QTableWidgetItem(emp["name"]))
            self.employee_table.setItem(row, 2, QTableWidgetItem(emp["designation"]))
            self.employee_table.setItem(row, 3, QTableWidgetItem("3"))
            self.employee_table.setItem(row, 4, QTableWidgetItem("PMTF"))
            division_dept = emp["division"] if emp["division"] else emp["department"]
            self.employee_table.setItem(row, 5, QTableWidgetItem(division_dept))
            self.employee_table.setItem(row, 6, QTableWidgetItem("Officers & Above"))
            self.employee_table.setItem(row, 7, QTableWidgetItem("Regular"))  
            self.employee_table.setItem(row, 8, QTableWidgetItem(str(emp["doj"])))
            rating_num = "3"  
            if emp["overall_rating"]:
                if "(5)" in emp["overall_rating"] or "Outstanding" in emp["overall_rating"]:
                    rating_num = "5"
                elif "(4)" in emp["overall_rating"] or "Exceeds" in emp["overall_rating"]:
                    rating_num = "4"
                elif "(3)" in emp["overall_rating"] or "Meets" in emp["overall_rating"]:
                    rating_num = "3"
                elif "(2)" in emp["overall_rating"] or "Below" in emp["overall_rating"]:
                    rating_num = "2"
                elif "(1)" in emp["overall_rating"] or "Serious" in emp["overall_rating"]:
                    rating_num = "1"
            self.employee_table.setItem(row, 9, QTableWidgetItem(rating_num))
            score_text = f"{emp['overall_percentage']:.1f}" if emp['overall_percentage'] else "0.0"
            self.employee_table.setItem(row, 10, QTableWidgetItem(score_text))
        self.employee_table.resizeColumnsToContents()

    def closeEvent(self, event):
        # logging.debug("Closing CurvedPerformanceView")
        if self.parent_app:
            self.parent_app.go_back()
        event.accept()