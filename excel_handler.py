import openpyxl
import os,sys
import json
from datetime import datetime
import logging

# Configure logging for debugging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ExcelHandler:
    def __init__(self):
        # Dynamically resolve the path for the assets folder
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            logging.debug(f"Running as executable, using sys._MEIPASS: {base_path}")
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))
            logging.debug(f"Running as script, using base path: {base_path}")

        self.file_path = os.path.join(base_path, "assets", "employee_performance_data.xlsx")
        self.config_path = os.path.join(base_path, "assets", "column_config.json")
        self.mandatory_columns = [
            "Employee ID", "Employee Name", "Department", "Designation",
            "Date of Joining", "Contract Expiry Date", "Division", "Exp in PMTF"
        ]
        self.visible_columns = []
        logging.debug(f"Excel file path set to: {self.file_path}")
        logging.debug(f"Config file path set to: {self.config_path}")

        # Check if file exists, create if not
        if not os.path.exists(self.file_path):
            logging.warning(f"Excel file not found at {self.file_path}. Creating a new one...")
            self.create_file()

        self.wb = None
        self.ws = None
        self.initialize_excel()
        self.load_column_config()


    def create_file(self):
        """Create a new Excel file with default headers if it doesn't exist."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance_Data"
        headers = [
            "Employee ID", "Employee Name", "Division", "Department", "Designation",
            "Date of Joining", "Exp in PMTF", "Date of Evaluation", "Contract Expiry Date",
            "Line Manager", "Entity Name",
            "KPI_1_Title", "KPI_1_Rating", "KPI_1_Weightage", "KPI_1_Weighted_Score",
            "KPI_2_Title", "KPI_2_Rating", "KPI_2_Weightage", "KPI_2_Weighted_Score",
            "KPI_3_Title", "KPI_3_Rating", "KPI_3_Weightage", "KPI_3_Weighted_Score",
            "KPI_4_Title", "KPI_4_Rating", "KPI_4_Weightage", "KPI_4_Weighted_Score",
            "KPI_5_Title", "KPI_5_Rating", "KPI_5_Weightage", "KPI_5_Weighted_Score",
            "KPI_6_Title", "KPI_6_Rating", "KPI_6_Weightage", "KPI_6_Weighted_Score",
            "Part_A_Total_Score",
            "Open_Clear_Communication_Rating", "Open_Clear_Communication_Weighted_Score",
            "Attitude_Team_Work_Collaboration_Rating", "Attitude_Team_Work_Collaboration_Weighted_Score",
            "Planning_Achievement_Focus_Rating", "Planning_Achievement_Focus_Weighted_Score",
            "Creativity_Initiatives_Rating", "Creativity_Initiatives_Weighted_Score",
            "Ownership_Self_Accountability_Rating", "Ownership_Self_Accountability_Weighted_Score",
            "Part_B_Total_Score",
            "Overall_Rating", "Overall_Percentage",
            "Promotion_Recommendation", "Retention_Recommendation",
            "Areas_of_Strength", "Areas_of_Development",
            "Comments", "Last_Updated",
            "Last_Year_Rating", "Last_Year_Increment",
            "Basic_Salary", "Gross_Amount", "Car_Allowance", "Fuel_Litre",
            "Fuel_Price", "House_Rent", "Medical", "Utilities", "Total_Salary",
            "Diff_Salary", "Diff_Conveyance", "Fuel_Litre_Adj", "Fuel_Price_Adj",
            "Diff_Car_Allowance", "Amount_Diff_Fuel",
            "Total_Salary_Adj", "Allowance_Adj", "Salary_Adj_Impact",
            "Salary_Increment_2425", "Training_Recommendations"
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        wb.save(self.file_path)
        logging.info(f"Created new Excel file at {self.file_path}")

    def initialize_excel(self):
        """Initialize the Excel workbook and worksheet."""
        try:
            self.wb = openpyxl.load_workbook(self.file_path)
            sheet_names = self.wb.sheetnames
            logging.debug(f"Available sheets: {sheet_names}")
            for sheet_name in ["Sheet1", "Performance_Data", "Data", "Employee_Data"]:
                if sheet_name in sheet_names:
                    self.ws = self.wb[sheet_name]
                    logging.info(f"Selected sheet: {sheet_name}")
                    break
            if not self.ws:
                raise ValueError(f"No valid sheet found in {self.file_path}. Available sheets: {sheet_names}")
        except Exception as e:
            logging.error(f"Failed to initialize Excel file: {str(e)}")
            raise Exception(f"Failed to initialize Excel file: {str(e)}")

    def load_column_config(self):
        """Load visible columns from config file."""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r') as f:
                    config = json.load(f)
                    self.visible_columns = self.mandatory_columns + [
                        col for col in config.get("visible_columns", []) if col not in self.mandatory_columns
                    ]
            else:
                headers = self.get_headers()
                self.visible_columns = headers  # All columns visible by default
                self.save_column_config()
            logging.info(f"Loaded visible columns: {self.visible_columns}")
        except Exception as e:
            logging.error(f"Error loading column config: {str(e)}")
            self.visible_columns = self.mandatory_columns  # Fallback to mandatory columns

    def save_column_config(self):
        """Save visible columns to config file."""
        try:
            config = {"visible_columns": [
                col for col in self.visible_columns if col not in self.mandatory_columns
            ]}
            with open(self.config_path, 'w') as f:
                json.dump(config, f, indent=4)
            logging.info(f"Saved column config: {config}")
        except Exception as e:
            logging.error(f"Error saving column config: {str(e)}")

    def get_headers(self):
        """Get all column headers from Excel file."""
        try:
            headers = [cell.value for cell in self.ws[1] if cell.value is not None]
            logging.info(f"Headers found: {headers}")
            return headers
        except Exception as e:
            logging.error(f"Error fetching headers: {str(e)}")
            raise Exception(f"Error fetching headers: {str(e)}")

    def get_visible_headers(self):
        """Get only visible column headers."""
        headers = self.get_headers()
        return [h for h in headers if h in self.visible_columns]

    def update_visible_columns(self, visible_columns):
        """Update visible columns list, keeping mandatory columns."""
        try:
            self.visible_columns = self.mandatory_columns + [
                col for col in visible_columns if col not in self.mandatory_columns
            ]
            self.save_column_config()
            logging.info(f"Updated visible columns: {self.visible_columns}")
        except Exception as e:
            logging.error(f"Error updating visible columns: {str(e)}")
            raise Exception(f"Error updating visible columns: {str(e)}")

    def add_column(self, column_name):
        """Add a new column to the Excel file."""
        try:
            headers = self.get_headers()
            if column_name in headers:
                return False, "Column already exists"
            if not column_name:
                return False, "Column name cannot be empty"
            self.ws.cell(row=1, column=len(headers)+1).value = column_name
            self.wb.save(self.file_path)
            if column_name not in self.visible_columns:
                self.visible_columns.append(column_name)
                self.save_column_config()
            logging.info(f"Added new column: {column_name}")
            return True, "Column added successfully"
        except Exception as e:
            logging.error(f"Error adding column: {str(e)}")
            return False, f"Error adding column: {str(e)}"

    def get_all_employees(self):
        """Retrieve all employees from the Excel file."""
        try:
            headers = self.get_headers()
            if not headers:
                raise ValueError("No headers found in Excel file")
            
            emp_id_col = None
            for i, header in enumerate(headers):
                header_str = str(header).strip().lower()
                if header_str in ["employee id", "emp id", "id", "employee_id"]:
                    emp_id_col = i
                    logging.info(f"Employee ID column found at index {i}: '{header}'")
                    break
            if emp_id_col is None:
                raise ValueError("Column 'Employee ID' not found in Excel file")
            
            employees = []
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and row[emp_id_col] and row[emp_id_col] != "":
                    employee = {}
                    for header in self.visible_columns:
                        if header in headers:
                            idx = headers.index(header)
                            employee[header] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
                    employees.append(employee)
            
            logging.info(f"Loaded {len(employees)} employees: {[emp['Employee ID'] for emp in employees]}")
            return employees
        except Exception as e:
            logging.error(f"Error fetching employees: {str(e)}")
            raise Exception(f"Error fetching employees: {str(e)}")

    def get_employee_data(self, emp_id):
        """Retrieve data for a specific employee by ID."""
        try:
            headers = self.get_headers()
            emp_id_col = None
            for i, header in enumerate(headers):
                if header and str(header).strip().lower() in ["employee id", "emp id", "id", "employee_id"]:
                    emp_id_col = i
                    break
            if emp_id_col is None:
                raise ValueError("Column 'Employee ID' not found in Excel file")
            
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and row[emp_id_col] == emp_id:
                    employee_data = {}
                    for header in self.visible_columns:
                        if header in headers:
                            idx = headers.index(header)
                            employee_data[header] = str(row[idx]) if idx < len(row) and row[idx] is not None else ""
                    logging.info(f"Employee data found for ID {emp_id}: {employee_data}")
                    return employee_data
            logging.warning(f"No employee found with ID {emp_id}")
            return None
        except Exception as e:
            logging.error(f"Error fetching employee data for ID {emp_id}: {str(e)}")
            raise Exception(f"Error fetching employee data for ID {emp_id}: {str(e)}")

    def save_employee_data(self, data):
        """Save or update employee data in the Excel file."""
        try:
            headers = self.get_headers()
            emp_id_col = None
            for i, header in enumerate(headers):
                if header and str(header).strip().lower() in ["employee id", "emp id", "id", "employee_id"]:
                    emp_id_col = i
                    break
            if emp_id_col is None:
                raise ValueError("Column 'Employee ID' not found in Excel file")
            
            emp_id = data.get("Employee ID")
            row_index = None
            for i, row in enumerate(self.ws.iter_rows(min_row=2), start=2):
                if row[emp_id_col].value == emp_id:
                    row_index = i
                    break
            
            if row_index:
                # Update existing row
                for col_idx, header in enumerate(headers):
                    if header in data:
                        self.ws.cell(row=row_index, column=col_idx + 1).value = data[header]
                logging.info(f"Updated employee data for ID {emp_id} at row {row_index}")
            else:
                # Append new row
                new_row = [data.get(header, "") for header in headers]
                self.ws.append(new_row)
                logging.info(f"Appended new employee data for ID {emp_id}")
            
            self.wb.save(self.file_path)
            logging.info(f"Excel file saved: {self.file_path}")
        except Exception as e:
            logging.error(f"Error saving employee data: {str(e)}")
            raise Exception(f"Error saving employee data: {str(e)}")

    def add_new_employee(self, emp_id, name, department, designation, joining_date, contract_expiry, division, exp_pmtf):
        """Add a new employee to the Excel file."""
        try:
            headers = self.get_headers()
            emp_id_col = None
            for i, header in enumerate(headers):
                if header and str(header).strip().lower() in ["employee id", "emp id", "id", "employee_id"]:
                    emp_id_col = i
                    break
            if emp_id_col is None:
                raise ValueError("Column 'Employee ID' not found in Excel file")
            
            # Check for duplicate Employee ID
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row and row[emp_id_col] == emp_id:
                    logging.warning(f"Duplicate Employee ID {emp_id} found")
                    return False, "Employee ID already exists"
            
            # Prepare new employee data
            new_employee = {
                "Employee ID": emp_id,
                "Employee Name": name,
                "Department": department,
                "Designation": designation,
                "Date of Joining": joining_date,
                "Contract Expiry Date": contract_expiry,
                "Division": division,
                "Exp in xyz": exp_pmtf,
                "Entity Name": "xyz",
                "Date of Evaluation": datetime.now().strftime("%Y-%m-%d")
            }
            
            # Append new row with default values for other columns
            row_data = [new_employee.get(header, "") for header in headers]
            self.ws.append(row_data)
            self.wb.save(self.file_path)
            logging.info(f"Added new employee {emp_id} to Excel file")
            return True, "Employee added successfully"
        except Exception as e:
            logging.error(f"Error adding new employee: {str(e)}")
            return False, f"Error adding new employee: {str(e)}"